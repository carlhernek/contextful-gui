"""File preview for the UI (spec section 4: preview method)."""
from __future__ import annotations

import base64
import csv
import io
import mimetypes
from pathlib import Path
from typing import Any

from contextful_sidecar.runtime.tools import _resolve

PREVIEW_CAP = 200_000
TABLE_ROW_CAP = 500
TEXT_EXTENSIONS = {
    ".txt", ".md", ".markdown", ".json", ".yaml", ".yml", ".py", ".js", ".ts", ".tsx",
    ".jsx", ".html", ".htm", ".css", ".xml", ".toml", ".ini", ".cfg", ".log",
    ".rs", ".go", ".java", ".sh", ".sql", ".rtf",
}
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".svg", ".bmp", ".ico"}
DOCX_EXTENSIONS = {".docx", ".doc"}


def _preview_text(data: str) -> dict[str, Any]:
    truncated = len(data) > PREVIEW_CAP
    if truncated:
        data = data[:PREVIEW_CAP] + "\n...[truncated]"
    return {"ok": True, "kind": "text", "content": data, "truncated": truncated}


def _preview_csv(data: str) -> dict[str, Any]:
    reader = csv.reader(io.StringIO(data))
    rows = list(reader)
    if not rows:
        return {"ok": True, "kind": "table", "table": {"headers": [], "rows": []}, "truncated": False}
    headers = rows[0]
    body = rows[1 : TABLE_ROW_CAP + 1]
    truncated = len(rows) - 1 > TABLE_ROW_CAP
    return {
        "ok": True,
        "kind": "table",
        "table": {"headers": headers, "rows": body},
        "truncated": truncated,
    }


def _preview_xlsx(target: Path) -> dict[str, Any]:
    from openpyxl import load_workbook

    wb = load_workbook(target, read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            return {"ok": False, "error": "empty workbook", "kind": "unsupported"}
        rows_iter = ws.iter_rows(values_only=True)
        first = next(rows_iter, None)
        if first is None:
            return {"ok": True, "kind": "table", "table": {"headers": [], "rows": []}, "truncated": False}
        headers = ["" if c is None else str(c) for c in first]
        body: list[list[str]] = []
        for row in rows_iter:
            body.append(["" if c is None else str(c) for c in row])
            if len(body) >= TABLE_ROW_CAP:
                break
        truncated = len(body) >= TABLE_ROW_CAP
        return {
            "ok": True,
            "kind": "table",
            "table": {"headers": headers, "rows": body},
            "truncated": truncated,
        }
    finally:
        wb.close()


def _preview_docx(target: Path) -> dict[str, Any]:
    from docx import Document

    doc = Document(target)
    lines = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    data = "\n".join(lines) if lines else "(empty document)"
    return _preview_text(data)


def _preview_image(target: Path) -> dict[str, Any]:
    raw = target.read_bytes()
    mime, _ = mimetypes.guess_type(target.name)
    if not mime or not mime.startswith("image/"):
        mime = "image/png"
    encoded = base64.b64encode(raw).decode("ascii")
    return {
        "ok": True,
        "kind": "image",
        "imageUrl": f"data:{mime};base64,{encoded}",
        "truncated": False,
    }


def preview_file(workspace: Path, path: str, base: str = "repos") -> dict[str, Any]:
    """Return a preview dict for a file under <base>/ in the workspace."""
    workspace = Path(workspace)
    rel = path if base in ("", ".") else f"{base.rstrip('/')}/{path.lstrip('/')}"
    try:
        target = _resolve(workspace, rel)
    except ValueError as exc:
        return {"ok": False, "error": str(exc), "path": rel, "kind": "unsupported"}
    if not target.exists():
        return {"ok": False, "error": "file not found", "path": rel, "kind": "unsupported"}
    if target.is_dir():
        return {"ok": False, "error": "path is a directory", "path": rel, "kind": "unsupported"}

    ext = target.suffix.lower()
    base_result: dict[str, Any] = {
        "path": rel,
        "name": target.name,
        "ext": target.suffix.lstrip("."),
        "size": target.stat().st_size,
    }

    if ext == ".csv":
        try:
            data = target.read_text(encoding="utf-8", errors="replace")
        except OSError as exc:
            return {**base_result, "ok": False, "error": str(exc), "kind": "unsupported"}
        return {**base_result, **_preview_csv(data)}

    if ext in (".xlsx", ".xlsm"):
        try:
            return {**base_result, **_preview_xlsx(target)}
        except Exception as exc:  # noqa: BLE001
            return {**base_result, "ok": False, "error": str(exc), "kind": "unsupported"}

    if ext in DOCX_EXTENSIONS:
        try:
            return {**base_result, **_preview_docx(target)}
        except Exception as exc:  # noqa: BLE001
            return {**base_result, "ok": False, "error": str(exc), "kind": "unsupported"}

    if ext in IMAGE_EXTENSIONS:
        try:
            return {**base_result, **_preview_image(target)}
        except OSError as exc:
            return {**base_result, "ok": False, "error": str(exc), "kind": "unsupported"}

    if ext == ".pdf":
        return {
            **base_result,
            "ok": True,
            "kind": "text",
            "content": f"PDF document ({base_result['size']} bytes). Open externally to view.",
            "truncated": False,
        }

    if ext not in TEXT_EXTENSIONS and ext != "":
        return {
            **base_result,
            "ok": False,
            "error": f"unsupported file type '{ext}'",
            "kind": "unsupported",
        }

    try:
        data = target.read_text(encoding="utf-8", errors="replace")
    except OSError as exc:
        return {**base_result, "ok": False, "error": str(exc), "kind": "unsupported"}
    return {**base_result, **_preview_text(data)}
